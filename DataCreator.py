from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By


class DataCreator:

    @staticmethod
    def create_full_information_sheet(filename):
        """
        Метод, создающий новую таблицу Excel с полными данными объекта
        """

        wb = Workbook()
        ws = wb.active

        data = [
            ["house_number", "road", "suburb", "city", "state", "region", "postcode", "country", "lat", "lon"],
            ["1/9 с8", "Кремлёвская набережная", "Тверской район", "Москва", "Москва", "Центральный федеральный округ",
             "119019", "Россия", "55.747624", "37.610703"]
        ]

        for row in data:
            ws.append(row)
        wb.save(filename)

    @staticmethod
    def init_table(filename):
        """
        Вспомогательный метод, создающий новую таблицу Excel с 3 заголовками
        """

        wb = Workbook()
        ws = wb.active
        ws.append(["name", "lat", "lon"])
        wb.save(filename)
        return wb

    @classmethod
    def create_landmarks_sheet(cls, filename):
        """
        Метод, берущий с сайта названия и координаты достопримечательностей Санкт-Петербурга и сохраняющий их в
        новую таблицу Excel
        """

        landmarks_url = "http://openarium.ru/Россия/Санкт-Петербург/Достопримечательности/"
        driver = webdriver.Chrome()

        driver.get(landmarks_url)
        landmarks = driver.find_elements(By.XPATH, "//section/h2")
        coords = driver.find_elements(By.XPATH, "//em")
        columns_count = len(landmarks)

        wb = cls.init_table(filename)
        ws = wb.active
        for i in range(2, columns_count):
            landmark = landmarks[i].text
            lat, lon = coords[i].text.split(',')
            ws.append([landmark, lat, lon])
        wb.save(filename)

    @staticmethod
    def create_queries_sheet(filename):
        """
        Метод, создающий новую таблицу Excel с фиксированными значениями для запроса и их координатами
        """

        data = [
            ["name", "lat", "lon"],
            ["Европолис Спб", 59.987307, 30.354264],
            ["Санкт-Петербургский Политехнический Университет Петра Великого", 60.007350, 30.373067],
            ["Место дуэли А. С. Пушкина", 59.995023, 30.302048],
            ["Петропавловская крепость", 59.950254, 30.316758],
            ["ТЮЗ Спб", 59.919858, 30.334794],
            ["Дом Зингера", 59.935575, 30.326008],
        ]

        wb = Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        wb.save(filename)
