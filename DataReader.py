import random
import string

from openpyxl import load_workbook


class DataReader:

    @staticmethod
    def generate_full_data(filename):
        """
        Метод, генерирующий из таблицы Excel словарь с полной информацией об объекте
        """
        wb = load_workbook(filename)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            yield {
                "address": {
                    "house_number": row[0],
                    "road": row[1],
                    "suburb": row[2],
                    "city": row[3],
                    "state": row[4],
                    "region": row[5],
                    "postcode": row[6],
                    "country": row[7]
                },
                "lat": row[8],
                "lon": row[9]
            }

    @staticmethod
    def generate_query_data(filename):
        """
        Метод, генерирующий из таблицы Excel информацию для нового запроса
        """

        wb = load_workbook(filename)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            yield {
                'q': row[0],
                'lat': row[1],
                'lon': row[2]
            }

    @staticmethod
    def generate_moscow_coords():
        """
        Метод, генерирующий случайные координаты Москвы
        """

        for i in range(10):
            yield [random.uniform(55.0, 56.0), random.uniform(36.0, 37.0)]

    @staticmethod
    def generate_incorrect_lats(count):
        """
        Метод, генерирующий координаты с неверным значением долготы
        """

        for i in range(count):
            if random.randint(0, 1):
                yield [random.uniform(-91.0, -180.0), random.uniform(1.0, 89.0)]
            else:
                yield [random.uniform(91.0, 180.0), random.uniform(1.0, 89.0)]

    @staticmethod
    def generate_incorrect_lons(count):
        """
        Метод, генерирующий координаты с неверным значением широты
        """
        for i in range(count):
            if random.randint(0, 1):
                yield [random.uniform(-100.0, -180.0), random.uniform(-100.0, -180.0)]
            else:
                yield [random.uniform(1.0, 89.0), random.uniform(91.0, 180.0)]

    @staticmethod
    def generate_incorrect_query(count):
        """
        Метод, генерирующий случайный текст запроса
        :return: str
        """
        for i in range(count):
            yield ''.join(random.sample(string.ascii_letters + string.digits, random.randint(1, 10)))
